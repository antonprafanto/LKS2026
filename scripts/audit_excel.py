"""Audit LKS2026-Penilaian-Juri.xlsx for formula and layout bugs."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "templates" / "LKS2026-Penilaian-Juri.xlsx"

BUGS: list[str] = []
WARNINGS: list[str] = []


def bug(msg: str) -> None:
    BUGS.append(msg)


def warn(msg: str) -> None:
    WARNINGS.append(msg)


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    sheet_names = set(wb.sheetnames)

    # --- Cross-sheet references ---
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                for ref in re.findall(r"(?:'([^']+)'|([A-Za-z][\w]*))!", v):
                    name = ref[0] or ref[1]
                    if name not in sheet_names:
                        bug(f"{ws.title}!{cell.coordinate}: referensi sheet tidak ada -> '{name}' dalam {v}")

    # --- Modul D bobot ---
    md = wb["Modul D"]
    bobot_sum = 0.0
    for r in range(7, 21):
        b = md.cell(r, 4).value
        if isinstance(b, (int, float)):
            bobot_sum += b
    total_label = md.cell(21, 4).value
    if abs(bobot_sum - 12.9) > 0.01:
        warn(f"Modul D: jumlah bobot item = {bobot_sum} (dokumen teknis: 12.9)")
    f21 = md.cell(21, 6).value
    if not (isinstance(f21, str) and "12/12.9" in f21):
        bug(f"Modul D F21 harus skala ke 12 poin, rumus={f21!r}")

    # --- Modul E bobot ---
    me = wb["Modul E"]
    i7 = me.cell(7, 9).value
    if i7 not in (60 / 9, "=60/9"):
        if not (isinstance(i7, str) and "60/9" in i7):
            warn(f"Modul E bobot kolom I: {i7!r} (harus =60/9)")

    # --- Rekapitulasi links ---
    rekap_expected = {
        6: ("Modul A", "H11"),
        7: ("Modul B", "I15"),
        8: ("Modul C", "H11"),
        9: ("Modul D", "F21"),
        10: ("Modul E", "J16"),
    }
    rekap = wb["Rekapitulasi"]
    for row, (sheet, coord) in rekap_expected.items():
        formula = rekap.cell(row, 4).value
        expected = f"='{sheet}'!{coord}"
        actual_cell = wb[sheet][coord].value
        if formula != expected:
            bug(f"Rekapitulasi D{row}: formula={formula!r}, expected {expected!r}")
        if not (isinstance(actual_cell, str) and actual_cell.startswith("=")):
            bug(f"{sheet}!{coord} bukan rumus total: {actual_cell!r}")

    # --- Modul E VLOOKUP ke Undian Kubus ---
    for r in range(7, 16):
        f = me.cell(r, 3).value
        if isinstance(f, str) and "Undian!" in f.replace("Undian Kubus", ""):
            bug(f"Modul E C{r}: referensi sheet salah (Undian tanpa Kubus)")
        if isinstance(f, str) and "Undian Kubus" not in f and "VLOOKUP" in f:
            bug(f"Modul E C{r}: VLOOKUP tanpa sheet Undian Kubus")
        for col, label in ((5, "Posisi Awal"), (6, "Slot Target"), (7, "OK Awal")):
            fx = me.cell(r, col).value
            if not (isinstance(fx, str) and fx.startswith("=IF")):
                bug(f"Modul E {label} baris {r} harus rumus IF (kolom {col})")
            if col == 6 and isinstance(fx, str) and "baris 24" not in fx and "baris 22" not in fx:
                bug(f"Modul E Slot Target baris {r} harus cek marker baris isian")
            if col == 7 and isinstance(fx, str) and ",7,FALSE)" not in fx.replace(" ", ""):
                bug(f"Modul E OK Awal baris {r} harus VLOOKUP Undian kolom G")

    # --- Undian layout constants ---
    und = wb["Undian Kubus"]
    if und.cell(10, 1).value != "K-M1":
        bug(f"Undian A10 harus K-M1, dapat {und.cell(10,1).value!r}")
    if und.cell(23, 1).value != "Rak1 · kiri":
        bug(f"Undian baris 23 header marker rusak: {und.cell(23,1).value!r}")
    if und.cell(27, 1).value != "ID Kubus":
        bug(f"Undian baris header posisi harus di 27, A27={und.cell(27,1).value!r}")
    if not str(und.cell(28, 2).value).startswith("=IF"):
        bug("Undian B28 harus rumus VLOOKUP")
    if not str(und.cell(28, 4).value).startswith("=IF"):
        bug("Undian D28 (Lokasi Awal) harus rumus otomatis")
    if not str(und.cell(28, 8).value).startswith("=IF"):
        bug("Undian H28 (Target Rak) harus rumus otomatis")
    if not str(und.cell(28, 9).value).startswith("=IF"):
        bug("Undian I28 (Target Marker) harus rumus otomatis")

    # --- Modul B total row ---
    mb = wb["Modul B"]
    if mb.cell(15, 2).value != "TOTAL MODUL B":
        bug(f"Modul B total label di B15 = {mb.cell(15,2).value!r}")
    if mb.cell(15, 9).value != "=I13+I14":
        bug(f"Modul B I15 = {mb.cell(15,9).value!r}")

    # --- Modul C total ---
    mc = wb["Modul C"]
    if mc.cell(11, 8).value != "=SUM(H7:H10)":
        bug(f"Modul C H11 = {mc.cell(11,8).value!r}")

    wb.close()

    print("=== AUDIT", XLSX.name, "===")
    print(f"BUGS: {len(BUGS)}")
    for b in BUGS:
        print("  [BUG]", b)
    print(f"WARNINGS: {len(WARNINGS)}")
    for w in WARNINGS:
        print("  [WARN]", w)
    if not BUGS and not WARNINGS:
        print("  Tidak ada masalah terdeteksi.")
    return 1 if BUGS else 0


if __name__ == "__main__":
    raise SystemExit(main())
