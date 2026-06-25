"""Deep audit for LKS2026 Excel workbooks (template + contoh)."""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"

MARKERS = ["Merah", "Hijau", "Biru", "Biru", "Merah", "Hijau", "Hijau", "Biru", "Merah"]
MARKER_VALUE_ROW = 24
EXPECTED_SHEETS = [
    "Panduan", "Modul A", "Modul B", "Modul C", "Modul D",
    "Undian Kubus", "Modul E", "Log Run Otonom", "Rekapitulasi",
]
CONTOH_POSITIONS = [
    ("K-M1", "Rak 1", "kiri"), ("K-M2", "Stand A", "depan"), ("K-M3", "Rak 2", "tengah"),
    ("K-H1", "Rak 3", "kiri"), ("K-H2", "Stand B", "depan"), ("K-H3", "Rak 1", "kanan"),
    ("K-B1", "Rak 2", "kiri"), ("K-B2", "Rak 3", "tengah"), ("K-B3", "Stand C", "depan"),
]
SLOT_IDX = {
    ("Rak 1", "kiri"): 0, ("Rak 1", "tengah"): 1, ("Rak 1", "kanan"): 2,
    ("Rak 2", "kiri"): 3, ("Rak 2", "tengah"): 4, ("Rak 2", "kanan"): 5,
    ("Rak 3", "kiri"): 6, ("Rak 3", "tengah"): 7, ("Rak 3", "kanan"): 8,
}
CATALOG = {
    "K-M1": ("Merah", "Lingkaran"), "K-M2": ("Merah", "Persegi"), "K-M3": ("Merah", "Segitiga"),
    "K-H1": ("Hijau", "Lingkaran"), "K-H2": ("Hijau", "Persegi"), "K-H3": ("Hijau", "Segitiga"),
    "K-B1": ("Biru", "Lingkaran"), "K-B2": ("Biru", "Persegi"), "K-B3": ("Biru", "Segitiga"),
}

BUGS: list[str] = []
WARNS: list[str] = []
INFOS: list[str] = []


def bug(file: str, msg: str) -> None:
    BUGS.append(f"[{file}] {msg}")


def warn(file: str, msg: str) -> None:
    WARNS.append(f"[{file}] {msg}")


def info(file: str, msg: str) -> None:
    INFOS.append(f"[{file}] {msg}")


def sudah_benar(rak: str, slot: str, warna: str, markers: list[str]) -> int:
    key = (rak, slot.lower())
    if key not in SLOT_IDX:
        return 0
    return 1 if markers[SLOT_IDX[key]] == warna else 0


def target_rak(warna: str, markers: list[str]) -> str:
    racks = []
    for lo, hi, name in [(0, 3, "Rak 1"), (3, 6, "Rak 2"), (6, 9, "Rak 3")]:
        if warna in markers[lo:hi]:
            racks.append(name)
    return ", ".join(racks)


def target_slots(warna: str, markers: list[str]) -> list[str]:
    labels = [
        "R1-S1 (kiri)", "R1-S2 (tengah)", "R1-S3 (kanan)",
        "R2-S1", "R2-S2", "R2-S3", "R3-S1", "R3-S2", "R3-S3",
    ]
    return [labels[i] for i, m in enumerate(markers) if m == warna]


def audit_workbook(path: Path, is_contoh: bool = False) -> None:
    name = path.name
    if not path.exists():
        bug(name, f"File tidak ditemukan: {path}")
        return

    wb = load_workbook(path, data_only=False)
    sheet_names = wb.sheetnames

    # --- struktur ---
    if "Sheet" in sheet_names:
        bug(name, "Sheet default tidak terhapus — bug generator lama")
    if "Panduan" not in sheet_names:
        bug(name, "Sheet Panduan hilang")
    for s in EXPECTED_SHEETS:
        if s not in sheet_names:
            bug(name, f"Sheet hilang: {s}")
    if len(sheet_names) != len(EXPECTED_SHEETS):
        warn(name, f"Jumlah sheet = {len(sheet_names)}, expected {len(EXPECTED_SHEETS)}: {sheet_names}")

    # --- referensi silang ---
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                if len(v) > 8000:
                    warn(name, f"{ws.title}!{cell.coordinate}: rumus sangat panjang ({len(v)} char)")
                for ref in re.findall(r"(?:'([^']+)'|([A-Za-z][\w]*))!", v):
                    ref_name = ref[0] or ref[1]
                    if ref_name not in sheet_names:
                        bug(name, f"{ws.title}!{cell.coordinate}: sheet '{ref_name}' tidak ada")

    und = wb["Undian Kubus"]
    me = wb["Modul E"]
    log = wb["Log Run Otonom"]
    rekap = wb["Rekapitulasi"]

    # --- layout Undian ---
    checks = [
        (10, 1, "K-M1", "katalog A10"),
        (18, 1, "K-B3", "katalog A18"),
        (23, 1, "Rak1 · kiri", "header marker A23"),
        (27, 1, "ID Kubus", "header posisi A27"),
        (27, 5, "Rak/Stand", "header E27"),
        (27, 8, "Target Rak", "header H27"),
    ]
    for r, c, expected, label in checks:
        actual = und.cell(r, c).value
        if actual != expected:
            bug(name, f"Undian {label}: expected {expected!r}, got {actual!r}")

    if not str(und.cell(28, 4).value).startswith("="):
        bug(name, "Undian D28 bukan rumus Lokasi Awal")
    if "TEXTJOIN" in str(und.cell(28, 8).value or ""):
        bug(name, "Undian H28 masih pakai TEXTJOIN — ganti rumus kompatibel Excel 2010+")
    if "COUNTA" not in str(und.cell(28, 8).value) and "COUNTA" not in str(und.cell(28, 9).value):
        warn(name, "Undian target tidak cek COUNTA baris marker kosong")

    # --- data validation Undian E/F ---
    dvs = und.data_validations.dataValidation
    if len(dvs) < 2:
        warn(name, f"Undian: data validation kurang ({len(dvs)}), dropdown Rak/Slot mungkin hilang")

    # --- Modul E struktur ---
    if me.cell(5, 1).value is None or "PETUNJUK" not in str(me.cell(5, 1).value):
        warn(name, "Modul E baris 5: petunjuk merah tidak ditemukan")

    for r in range(7, 16):
        if me.cell(r, 9).value != "=60/9":
            bug(name, f"Modul E I{r} bobot harus =60/9, got {me.cell(r, 9).value!r}")
        skor = me.cell(r, 10).value
        if skor != f'=IF(H{r}="",0,H{r}*I{r})':
            bug(name, f"Modul E J{r} rumus skor salah: {skor!r}")
        b_formula = me.cell(r, 2).value
        if not (isinstance(b_formula, str) and "Undian Kubus" in b_formula and f"A{21 + r}" in b_formula):
            bug(name, f"Modul E B{r} harus tarik dari Undian A{21 + r}")

    total_j = me.cell(16, 10).value
    if total_j != "=ROUND(SUM(J7:J15),2)":
        bug(name, f"Modul E J16 total salah: {total_j!r}")

    # --- Log Run -> Modul E ---
    if "Modul E" not in str(log.cell(log.max_row - 20, 3).value if False else ""):
        pass
    log_count = log.cell(35, 3).value if log.max_row >= 35 else None
    for cell_ref, expected_fragment in [
        (log["C35"] if log.max_row >= 35 else None, "COUNTIF"),
    ]:
        pass
    # find hasil singkat rows
    hasil_found = False
    for row in log.iter_rows(min_row=1, max_row=log.max_row):
        if row[0].value and "Kubus berhasil" in str(row[0].value):
            hasil_found = True
            count_formula = row[2].value
            if not (isinstance(count_formula, str) and "Modul E" in count_formula and "H7:H15" in count_formula):
                bug(name, f"Log Run COUNTIF kubus: {count_formula!r}")
            skor_row = row[0].row + 1
            skor_f = log.cell(skor_row, 3).value
            if skor_f != "='Modul E'!J16":
                bug(name, f"Log Run skor Modul E: {skor_f!r}")
    if not hasil_found:
        warn(name, "Log Run: bagian HASIL SINGKAT tidak ditemukan")

    # --- Rekapitulasi ---
    if rekap.cell(10, 4).value != "='Modul E'!J16":
        bug(name, f"Rekapitulasi D10: {rekap.cell(10,4).value!r}")
    if rekap.cell(11, 4).value != "=SUM(D6:D10)":
        bug(name, f"Rekapitulasi total: {rekap.cell(11,4).value!r}")

    # --- Modul D skala ---
    md = wb["Modul D"]
    if "12/12.9" not in str(md.cell(21, 6).value):
        bug(name, "Modul D F21 tidak skala 12/12.9")

    # --- isi CONTOH ---
    if is_contoh:
        markers = [und.cell(MARKER_VALUE_ROW, c).value for c in range(1, 10)]
        if markers != MARKERS:
            bug(name, f"Baris {MARKER_VALUE_ROW} marker contoh salah: {markers}")

        for i, (kid, rak, slot) in enumerate(CONTOH_POSITIONS):
            r = 28 + i
            if und.cell(r, 1).value != kid:
                bug(name, f"Undian A{r}: expected {kid}, got {und.cell(r,1).value!r}")
            if und.cell(r, 5).value != rak:
                bug(name, f"Undian E{r}: expected {rak}, got {und.cell(r,5).value!r}")
            if und.cell(r, 6).value != slot:
                bug(name, f"Undian F{r}: expected {slot}, got {und.cell(r,6).value!r}")

            warna, bentuk = CATALOG[kid]
            lokasi = f"{rak} — {slot}"
            tr = target_rak(warna, MARKERS)
            slots = target_slots(warna, MARKERS)
            sb = sudah_benar(rak, slot, warna, MARKERS)

            # Modul E row 7+i
            er = 7 + i
            if me.cell(er, 2).value != kid:
                if not (isinstance(me.cell(er, 2).value, str) and me.cell(er, 2).value.startswith("=IF")):
                    bug(name, f"Modul E B{er}: expected formula ke Undian A{28+i}")
            elif und.cell(28 + i, 1).value != kid:
                bug(name, f"Undian A{28+i}={und.cell(28+i,1).value!r} tapi Modul E expect {kid}")
            # Warna/Bentuk are formulas - check formula exists
            if not str(me.cell(er, 3).value).startswith("=IF"):
                bug(name, f"Modul E C{er} bukan rumus warna")

            # Hasil filled in contoh
            hasil = me.cell(er, 8).value
            if hasil is None or hasil == "":
                warn(name, f"Modul E H{er} hasil contoh kosong")

        # Skor total contoh: 8 sukses * 60/9
        hasil_vals = [me.cell(7 + i, 8).value for i in range(9)]
        if hasil_vals.count(1) != 8:
            warn(name, f"Contoh: {hasil_vals.count(1)} sukses (dokumen bilang 8): {hasil_vals}")

    else:
        # template kosong: baris 22 harus kosong
        for c in range(1, 10):
            if und.cell(MARKER_VALUE_ROW, c).value is not None:
                warn(name, f"Template kosong tapi baris {MARKER_VALUE_ROW} col{c} punya data: {und.cell(MARKER_VALUE_ROW,c).value!r}")

    # --- UX / kompatibilitas ---
    if "TEXTJOIN" in str(me.cell(7, 6).value or ""):
        info(name, "Modul E masih pakai TEXTJOIN — perlu Excel 2019+")
    if "→" in str(me.cell(7, 5).value) or "isi" in str(me.cell(7, 5).value).lower():
        info(name, "Modul E menampilkan petunjuk jika Undian belum lengkap (by design)")

    # --- skenario vs Sudah Benar (hanya file contoh) ---
    if is_contoh:
        expected_sb = {
            ("K-M1", "Rak 1", "kiri"): 1,
            ("K-M2", "Stand A", "depan"): 0,
            ("K-M3", "Rak 2", "tengah"): 1,
            ("K-H1", "Rak 3", "kiri"): 1,
            ("K-H2", "Stand B", "depan"): 0,
            ("K-H3", "Rak 1", "kanan"): 0,
            ("K-B1", "Rak 2", "kiri"): 1,
            ("K-B2", "Rak 3", "tengah"): 1,  # marker Biru = warna kubus (posisi benar)
            ("K-B3", "Stand C", "depan"): 0,
        }
        for kid, rak, slot in CONTOH_POSITIONS:
            warna = CATALOG[kid][0]
            sb = sudah_benar(rak, slot, warna, MARKERS)
            exp = expected_sb.get((kid, rak, slot))
            if exp is not None and sb != exp:
                warn(name, f"Sudah Benar {kid}@{rak}/{slot}: simulasi {sb}, skenario doc ~{exp}")

    wb.close()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    files = [
        (TEMPLATES / "LKS2026-Penilaian-Juri.xlsx", False),
        (TEMPLATES / "LKS2026-Penilaian-Juri-CONTOH.xlsx", True),
    ]
    extra = TEMPLATES / "LKS2026-Penilaian-Juri-NEW.xlsx"
    if extra.exists():
        warn("repo", f"File duplikat lama ada: {extra.name} — bisa dihapus")

    for path, is_contoh in files:
        audit_workbook(path, is_contoh=is_contoh)

    print("=" * 60)
    print("AUDIT MENDALAM EXCEL LKS 2026")
    print("=" * 60)
    print(f"\nBUGS ({len(BUGS)}):")
    for b in BUGS:
        print(" ", b)
    print(f"\nWARNINGS ({len(WARNS)}):")
    for w in WARNS:
        print(" ", w)
    print(f"\nINFO ({len(INFOS)}):")
    for i in INFOS:
        print(" ", i)
    if not BUGS and not WARNS:
        print("\nSemua cek struktural lulus.")
    print(f"\nRingkasan: {len(BUGS)} bug | {len(WARNS)} warning | {len(INFOS)} info")
    return 1 if BUGS else 0


if __name__ == "__main__":
    raise SystemExit(main())
