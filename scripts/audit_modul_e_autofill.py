"""Audit Modul E auto-fill: Undian Kubus formulas and doc sync."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "templates" / "LKS2026-Penilaian-Juri.xlsx"

BUGS: list[str] = []
WARNS: list[str] = []

MARKERS = ["Merah", "Hijau", "Biru", "Biru", "Merah", "Hijau", "Hijau", "Biru", "Merah"]
MARKER_VALUE_ROW = 24
SLOT_LABELS = [
    "R1-S1 (kiri)", "R1-S2 (tengah)", "R1-S3 (kanan)",
    "R2-S1", "R2-S2", "R2-S3", "R3-S1", "R3-S2", "R3-S3",
]
CATALOG = {
    "K-M1": "Merah", "K-M2": "Merah", "K-M3": "Merah",
    "K-H1": "Hijau", "K-H2": "Hijau", "K-H3": "Hijau",
    "K-B1": "Biru", "K-B2": "Biru", "K-B3": "Biru",
}
SLOT_IDX = {
    ("Rak 1", "kiri"): 0, ("Rak 1", "s1"): 0, ("Rak 1", "r1-s1"): 0,
    ("Rak 1", "tengah"): 1, ("Rak 1", "s2"): 1, ("Rak 1", "r1-s2"): 1,
    ("Rak 1", "kanan"): 2, ("Rak 1", "s3"): 2, ("Rak 1", "r1-s3"): 2,
    ("Rak 2", "kiri"): 3, ("Rak 2", "s1"): 3, ("Rak 2", "r2-s1"): 3,
    ("Rak 2", "tengah"): 4, ("Rak 2", "s2"): 4, ("Rak 2", "r2-s2"): 4,
    ("Rak 2", "kanan"): 5, ("Rak 2", "s3"): 5, ("Rak 2", "r2-s3"): 5,
    ("Rak 3", "kiri"): 6, ("Rak 3", "s1"): 6, ("Rak 3", "r3-s1"): 6,
    ("Rak 3", "tengah"): 7, ("Rak 3", "s2"): 7, ("Rak 3", "r3-s2"): 7,
    ("Rak 3", "kanan"): 8, ("Rak 3", "s3"): 8, ("Rak 3", "r3-s3"): 8,
}


def sudah_benar(rak: str, slot: str, warna: str) -> int | str:
    if not rak or not slot:
        return ""
    key = (rak, slot.lower())
    if key not in SLOT_IDX:
        return 0
    return 1 if MARKERS[SLOT_IDX[key]] == warna else 0


def target_rak(warna: str) -> str:
    racks = []
    for lo, hi, name in [(0, 3, "Rak 1"), (3, 6, "Rak 2"), (6, 9, "Rak 3")]:
        if warna in MARKERS[lo:hi]:
            racks.append(name)
    return ", ".join(racks)


def main() -> int:
    wb = load_workbook(XLSX, data_only=False)
    und = wb["Undian Kubus"]
    me = wb["Modul E"]

    if und.cell(23, 1).value != "Rak1 · kiri":
        BUGS.append(f"Baris 23 harus header Rak1·kiri, dapat {und.cell(23,1).value!r}")
    if und.cell(MARKER_VALUE_ROW, 1).value is not None and str(und.cell(MARKER_VALUE_ROW, 1).value).startswith("="):
        BUGS.append(f"Baris {MARKER_VALUE_ROW} harus kosong untuk isian warna marker (bukan rumus)")
    if und.cell(27, 1).value != "ID Kubus":
        BUGS.append("Header posisi kubus harus di baris 27")

    for r in range(28, 37):
        for col, name in ((4, "Lokasi Awal"), (7, "Sudah Benar"), (8, "Target Rak"), (9, "Target Marker")):
            v = und.cell(r, col).value
            if not (isinstance(v, str) and v.startswith("=")):
                BUGS.append(f"Undian {name} baris {r} bukan rumus")

    for r in range(7, 16):
        for col, vcol in ((5, 5), (6, None), (7, None)):
            f = me.cell(r, col).value
            if not (isinstance(f, str) and f.startswith("=IF")):
                BUGS.append(f"Modul E baris {r} kolom {col}: bukan rumus petunjuk/otomatis")
            if col == 5 and isinstance(f, str) and "Undian Kubus" not in f:
                BUGS.append(f"Modul E Posisi Awal baris {r} harus referensi Undian")
            if col == 6 and isinstance(f, str) and f"baris {MARKER_VALUE_ROW}" not in f:
                BUGS.append(f"Modul E Slot Target baris {r} harus cek marker baris {MARKER_VALUE_ROW}")
            if col == 7 and isinstance(f, str) and ",7,FALSE)" not in f.replace(" ", ""):
                BUGS.append(f"Modul E OK Awal baris {r} harus VLOOKUP Undian kolom G")

    # Logika Sudah Benar
    cases = [
        ("K-M1", "Rak 1", "kiri", 1),
        ("K-M2", "Stand A", "depan", 0),
        ("K-H1", "Rak 3", "kiri", 1),
        ("K-B2", "Rak 3", "tengah", 1),
        ("K-M3", "Rak 2", "tengah", 1),  # marker R2-S2 = Merah
    ]
    for kid, rak, slot, expected in cases:
        got = sudah_benar(rak, slot, CATALOG[kid])
        if got != expected:
            BUGS.append(f"Sudah Benar {kid} @ {rak}/{slot}: expected {expected}, got {got}")

    # Target Rak: Merah ada di 3 rak dengan marker skenario
    tr = target_rak("Merah")
    if tr != "Rak 1, Rak 2, Rak 3":
        WARNS.append(f"Target Rak Merah = {tr!r} (bisa membingungkan juri — banyak opsi)")

    # TEXTJOIN tidak dipakai lagi (kompatibilitas Excel 2010+)
    if "TEXTJOIN" in str(und.cell(25, 8).value or ""):
        BUGS.append("Undian H25 masih pakai TEXTJOIN")

    # Kolom G formula only — tidak bisa override manual
    WARNS.append("Kolom Sudah Benar (G) hanya rumus — juri tidak bisa override manual")

    # Stand / slot bebas tidak terdeteksi benar
    WARNS.append("Posisi di Stand (bukan Rak 1-3) selalu Sudah Benar=0 — sesuai desain")

    # Doc sync
    ppt = (ROOT / "scripts" / "generate_briefing_slides.py").read_text(encoding="utf-8")
    if "baris 24" not in ppt.lower() and "baris 22" not in ppt.lower():
        WARNS.append("Briefing PPT belum menyebut isian marker (baris 24)")

    skenario = (ROOT / "docs" / "SKENARIO-SOAL-CONTOH.md").read_text(encoding="utf-8")
    if "baris 22" not in skenario.lower():
        WARNS.append("SKENARIO-SOAL-CONTOH belum update alur Excel otomatis")

    wb.close()

    print("=== AUDIT Modul E Auto-Fill ===")
    for b in BUGS:
        print("[BUG]", b)
    for w in WARNS:
        print("[WARN]", w)
    if not BUGS and not WARNS:
        print("Tidak ada masalah terdeteksi.")
    else:
        print(f"\nTotal: {len(BUGS)} bug, {len(WARNS)} warning")
    return 1 if BUGS else 0


if __name__ == "__main__":
    raise SystemExit(main())
