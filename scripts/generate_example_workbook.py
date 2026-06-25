"""Generate a pre-filled example workbook so judges can see Modul E auto-fill."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

import generate_excel_templates as gen

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "templates" / "LKS2026-Penilaian-Juri.xlsx"
OUT = ROOT / "templates" / "LKS2026-Penilaian-Juri-CONTOH.xlsx"
OUT_XLSM = OUT.with_suffix(".xlsm")

# Marker baris 24 — sesuai skenario latihan (Rak1:M,H,B | Rak2:B,M,H | Rak3:H,B,M)
MARKERS = ["Merah", "Hijau", "Biru", "Biru", "Merah", "Hijau", "Hijau", "Biru", "Merah"]

# Contoh posisi obstacle (skenario latihan — 6 dari 10 dipakai)
OBSTACLE_EXAMPLE = [
    "kiri GATE",
    "tengah kiri arena",
    "tengah kanan arena",
    "kanan atas Rak 2",
    "bawah tengah",
    "dekat Rak 3",
    "",
    "",
    "",
    "",
]

# (ID kubus, Rak/Stand, Slot)
POSITIONS = [
    ("K-M1", "Rak 1", "kiri"),
    ("K-M2", "Stand A", "depan"),
    ("K-M3", "Rak 2", "tengah"),
    ("K-H1", "Rak 3", "kiri"),
    ("K-H2", "Stand B", "depan"),
    ("K-H3", "Rak 1", "kanan"),
    ("K-B1", "Rak 2", "kiri"),
    ("K-B2", "Rak 3", "tengah"),
    ("K-B3", "Stand C", "depan"),
]


def fill_example(wb_path: Path) -> None:
    wb = load_workbook(wb_path)
    und = wb["Undian Kubus"]
    me = wb["Modul E"]
    log = wb["Log Run Otonom"]

    und["B3"] = "RUN-01"
    und["E3"] = "05"
    und["H3"] = "SMK Contoh"
    und["B4"] = "H2"
    und["E4"] = "24/06/2026"
    und["H4"] = "2"
    und["B5"] = "Trial"

    for col, warna in enumerate(MARKERS, 1):
        und.cell(gen.MARKER_VALUE_ROW, col, warna)

    for i, (kid, rak, slot) in enumerate(POSITIONS):
        r = gen.UNDIAN_POS_FIRST + i
        und.cell(r, 1, kid)
        und.cell(r, 5, rak)
        und.cell(r, 6, slot)

    # Baris obstacle — cari baris data (teks "Obs 1" di header)
    obs_data_row = None
    for row in range(gen.UNDIAN_POS_LAST + 1, gen.UNDIAN_POS_LAST + 8):
        if und.cell(row, 1).value == "Obs 1":
            obs_data_row = row + 1
            break
    if obs_data_row:
        for c, val in enumerate(OBSTACLE_EXAMPLE, 1):
            if val:
                und.cell(obs_data_row, c, val)

    me["B3"] = "05"
    me["D3"] = "H2"
    me["H3"] = "Trial"
    me["B4"] = "2"
    me["D4"] = "08:42"

    # Hasil contoh run 1 (8 sukses dari skenario)
    hasil = [1, 1, 1, 1, 0, 1, 1, 1, 1]
    for i, val in enumerate(hasil):
        me.cell(7 + i, 8, val)

    log["B3"] = "05"
    log["E7"] = "RUN-01"
    log["H7"] = "Trial"
    log["B8"] = "08:40:00"
    log["E8"] = "08:42:00"
    log["H8"] = "08:55:00"

    wb.save(wb_path)
    wb.close()


def main() -> None:
    if not TEMPLATE.exists():
        gen.main()
    shutil.copy2(TEMPLATE, OUT)
    fill_example(OUT)
    print(f"Created example workbook: {OUT}")

    try:
        from add_log_run_stopwatch import apply_stopwatch

        apply_stopwatch(OUT, OUT_XLSM)
        print(f"Created example workbook: {OUT_XLSM} (stopwatch)")
    except Exception as exc:
        print(f"WARN: CONTOH .xlsm not created — {exc}")

    print("Buka file .xlsm di Excel - sheet Log Run Otonom - tombol stopwatch.")


if __name__ == "__main__":
    main()
