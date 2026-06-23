"""Generate Excel scoring templates with auto formulas for LKS 2026 judges."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "templates" / "LKS2026-Penilaian-Juri.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATALOG = [
    ("K-M1", "Merah", "Lingkaran", "255-0-0"),
    ("K-M2", "Merah", "Persegi", "255-0-0"),
    ("K-M3", "Merah", "Segitiga", "255-0-0"),
    ("K-H1", "Hijau", "Lingkaran", "0-255-0"),
    ("K-H2", "Hijau", "Persegi", "0-255-0"),
    ("K-H3", "Hijau", "Segitiga", "0-255-0"),
    ("K-B1", "Biru", "Lingkaran", "0-0-255"),
    ("K-B2", "Biru", "Persegi", "0-0-255"),
    ("K-B3", "Biru", "Segitiga", "0-0-255"),
]

MARKER_LABELS = [
    "R1-S1 (kiri)", "R1-S2 (tengah)", "R1-S3 (kanan)",
    "R2-S1", "R2-S2", "R2-S3",
    "R3-S1", "R3-S2", "R3-S3",
]

# Baris tetap di sheet Undian Kubus (untuk rumus antar-sheet)
UNDIAN_CAT_FIRST = 10
UNDIAN_CAT_LAST = 18
UNDIAN_POS_FIRST = 25
UNDIAN_POS_LAST = 33
UNDIAN_SHEET = "'Undian Kubus'"
CAT_RANGE = f"{UNDIAN_SHEET}!$A${UNDIAN_CAT_FIRST}:$D${UNDIAN_CAT_LAST}"
POS_RANGE = f"{UNDIAN_SHEET}!$A${UNDIAN_POS_FIRST}:$E${UNDIAN_POS_LAST}"
MODUL_D_BOBOT_SUM = 12.9  # sesuai dokumen teknis (D2a 0.3 + D2b 0.6)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def sheet_title(ws, text: str, merge_end: str = "K1") -> None:
    ws["A1"] = text
    ws.merge_cells(f"A1:{merge_end}")
    ws["A1"].font = Font(bold=True, size=14)


def sheet_undian_kubus(wb: Workbook) -> None:
    ws = wb.create_sheet("Undian Kubus")
    sheet_title(ws, "LEMBAR UNDIAN POSISI KUBUS — LKS 2026")

    fields = [
        ("A3", "Run ID:", "B3"), ("D3", "Tim No:", "E3"), ("G3", "Nama Tim:", "H3"),
        ("A4", "Hari:", "B4"), ("D4", "Tanggal:", "E4"), ("G4", "Arena No:", "H4"),
        ("A5", "Fase:", "B5"), ("D5", "Juru Acak:", "E5"), ("G5", "Juri Saksi:", "H5"),
        ("A6", "Waktu Undian:", "B6"),
    ]
    for label_cell, label, value_cell in fields:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)

    ws["A8"] = "KATALOG KUBUS (referensi — 9 buah)"
    ws["A8"].font = Font(bold=True, size=12)
    cat_hdr = 9
    for c, h in enumerate(["ID Kubus", "Warna", "Bentuk", "RGB"], 1):
        ws.cell(cat_hdr, c, h)
    style_header_row(ws, cat_hdr, 4)
    for i, row in enumerate(CATALOG, UNDIAN_CAT_FIRST):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)

    cat_range = CAT_RANGE

    m_title = UNDIAN_CAT_LAST + 2
    ws.cell(m_title, 1, "KONFIGURASI MARKER RAK (isi warna: Merah / Hijau / Biru)")
    ws.cell(m_title, 1).font = Font(bold=True, size=12)
    m_hdr = m_title + 1
    for c, lbl in enumerate(MARKER_LABELS, 1):
        ws.cell(m_hdr, c, lbl)
    style_header_row(ws, m_hdr, len(MARKER_LABELS))

    p_title = m_hdr + 2
    ws.cell(p_title, 1, "POSISI AWAL SETELAH UNDIAN (isi ID kubus — kolom lain otomatis)")
    ws.cell(p_title, 1).font = Font(bold=True, size=12)
    pos_hdr = UNDIAN_POS_FIRST - 1
    pos_headers = [
        "ID Kubus", "Warna", "Bentuk", "Lokasi Awal", "Rak/Stand", "Slot",
        "Sudah Benar (1/0)", "Target Rak", "Target Marker", "Perlu Dipindah?", "Catatan",
    ]
    for c, h in enumerate(pos_headers, 1):
        ws.cell(pos_hdr, c, h)
    style_header_row(ws, pos_hdr, len(pos_headers))

    for r in range(UNDIAN_POS_FIRST, UNDIAN_POS_LAST + 1):
        ws.cell(r, 2, f'=IF(A{r}="","",VLOOKUP(A{r},{cat_range},2,FALSE))')
        ws.cell(r, 3, f'=IF(A{r}="","",VLOOKUP(A{r},{cat_range},3,FALSE))')
        ws.cell(r, 10, f'=IF(G{r}="","",IF(G{r}=1,0,1))')

    obs_title = UNDIAN_POS_LAST + 2
    ws.cell(obs_title, 1, "OBSTACLE (catat area / deskripsi posisi)")
    ws.cell(obs_title, 1).font = Font(bold=True)
    obs_hdr = obs_title + 1
    for c in range(1, 11):
        ws.cell(obs_hdr, c, f"Obs {c}")
    style_header_row(ws, obs_hdr, 10)

    sig = obs_hdr + 2
    ws.cell(sig, 1, "Tanda tangan Juru Acak:")
    ws.cell(sig, 4, "Tanda tangan Juri:")
    ws.cell(sig, 7, "Tanda tangan Peserta (saksi):")

    set_widths(ws, {
        1: 12, 2: 10, 3: 10, 4: 14, 5: 12, 6: 8, 7: 14, 8: 12, 9: 14, 10: 14, 11: 18,
    })


def sheet_log_run(wb: Workbook) -> None:
    ws = wb.create_sheet("Log Run Otonom")
    sheet_title(ws, "LEMBAR LOG RUN OTONOM — LKS 2026")

    ws["A3"], ws["C3"], ws["E3"], ws["G3"] = "Tim No:", "Nama Tim:", "Kontingen:", "Arena No:"
    ws["A4"], ws["C4"], ws["E4"], ws["G4"] = "Hari:", "Tanggal:", "Juri Penilai:", "Juri Pendamping:"

    ws["A6"] = "DATA RUN"
    ws["A6"].font = Font(bold=True)
    run_fields = [
        ("A7", "Run No:", "B7"), ("D7", "Run ID:", "E7"), ("G7", "Trial/Final:", "H7"),
        ("A8", "Waktu SIAP:", "B8"), ("D8", "Waktu START:", "E8"), ("G8", "Waktu Selesai:", "H8"),
        ("A9", "Undian ulang (1/0):", "B9"), ("D9", "Retry (1/0):", "E9"), ("G9", "Retry No:", "H9"),
    ]
    for label_cell, label, val_cell in run_fields:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)

    ws["A11"] = "CHECKLIST PROSEDUR JURI (isi 1 = selesai)"
    ws["A11"].font = Font(bold=True)
    checks = [
        "Layout arena dijelaskan ke peserta",
        "Peserta nyatakan SIAP",
        "Laptop tidak disentuh setelah SIAP",
        "Posisi kubus diacak / diacak ulang",
        "Sinyal START diberikan",
        "Peserta tekan tombol START di robot",
        "Run berjalan otonom",
        "Hasil dicatat di sheet Modul E",
    ]
    chk_hdr = 12
    ws.cell(chk_hdr, 1, "No")
    ws.cell(chk_hdr, 2, "Item")
    ws.cell(chk_hdr, 3, "OK (1/0)")
    style_header_row(ws, chk_hdr, 3)
    for i, item in enumerate(checks, chk_hdr + 1):
        ws.cell(i, 1, i - chk_hdr)
        ws.cell(i, 2, item)
        ws.cell(i, 3, "")

    ws.cell(chk_hdr + len(checks) + 2, 1, "Checklist selesai:")
    done_row = chk_hdr + len(checks) + 2
    ws.cell(done_row, 2, f"=IF(COUNTIF(C{chk_hdr+1}:C{chk_hdr+len(checks)},1)=8,\"LENGKAP\",\"BELUM\")")

    krono_hdr = done_row + 2
    ws.cell(krono_hdr, 1, "KRONOLOGI KEJADIAN")
    ws.cell(krono_hdr, 1).font = Font(bold=True)
    kh = krono_hdr + 1
    for c, h in enumerate(["Waktu", "Kejadian", "Catatan"], 1):
        ws.cell(kh, c, h)
    style_header_row(ws, kh, 3)
    defaults = [
        ("", "Robot start dari posisi start", ""),
        ("", "Navigasi / pick / place", ""),
        ("", "Robot selesai / timeout / STOP", ""),
    ]
    for i, row in enumerate(defaults, kh + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    for r in range(kh + 4, kh + 9):
        ws.cell(r, 1, "")

    hasil = kh + 10
    ws.cell(hasil, 1, "HASIL SINGKAT")
    ws.cell(hasil, 1).font = Font(bold=True)
    ws.cell(hasil + 1, 1, "Kubus berhasil ditempatkan benar:")
    ws.cell(hasil + 1, 3, "=COUNTIF('Modul E'!H7:H15,1)")
    ws.cell(hasil + 1, 4, "/ 9")
    ws.cell(hasil + 2, 1, "Skor Modul E run ini:")
    ws.cell(hasil + 2, 3, "='Modul E'!J16")
    ws.cell(hasil + 2, 4, "/ 60")
    ws.cell(hasil + 3, 1, "Pelanggaran (ya/tidak):")
    ws.cell(hasil + 3, 3, "")
    ws.cell(hasil + 4, 1, "Jenis pelanggaran:")
    ws.cell(hasil + 4, 3, "")

    sig = hasil + 6
    ws.cell(sig, 1, "Tanda tangan Juri:")
    ws.cell(sig, 4, "Tanda tangan Peserta:")
    ws.cell(sig, 7, "Tanda tangan Chief Expert:")

    set_widths(ws, {1: 28, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14, 7: 18, 8: 14})


def sheet_modul_a(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Modul A"
    ws["A1"] = "MODUL A — Organisasi & Manajemen Kerja (8 poin, Judgement)"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"], ws["G3"] = "Tim No:", "Nama Tim:", "Kontingen:", "Tanggal:"
    ws["A4"], ws["C4"], ws["E4"] = "Juri 1:", "Juri 2:", "Juri 3:"

    headers = [
        "Sub", "Kriteria", "Max", "Juri 1", "Juri 2", "Juri 3",
        "Rata-rata", "Skor Akhir", "Selisih", "Catatan",
    ]
    start = 6
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(headers))

    rows = [
        ("A1", "Efisiensi Penjadwalan Kerja", 2.0),
        ("A2", "Kolaborasi Tim & Manajemen Tugas", 2.0),
        ("A3", "Organisasi & Kerapian Area Kerja (5S)", 2.0),
        ("A4", "Restorasi Area Kerja", 2.0),
    ]
    for idx, (sub, name, mx) in enumerate(rows, start + 1):
        r = idx
        ws.cell(r, 1, sub)
        ws.cell(r, 2, name)
        ws.cell(r, 3, mx)
        ws.cell(r, 7, f"=IF(COUNT(D{r}:F{r})=0,\"\",AVERAGE(D{r}:F{r}))")
        ws.cell(r, 8, f"=IF(G{r}=\"\",0,G{r}*C{r}/3)")
        ws.cell(r, 9, f"=IF(COUNT(D{r}:F{r})<2,\"\",MAX(D{r}:F{r})-MIN(D{r}:F{r}))")
        for c in range(4, 7):
            ws.cell(r, c).number_format = "0.0"

    total_row = start + 1 + len(rows)
    ws.cell(total_row, 2, "TOTAL MODUL A")
    ws.cell(total_row, 3, 8.0)
    ws.cell(total_row, 8, f"=SUM(H{start+1}:H{total_row-1})")
    ws.cell(total_row, 2).font = Font(bold=True)

    set_widths(ws, {1: 6, 2: 38, 3: 8, 4: 10, 5: 10, 6: 10, 7: 12, 8: 12, 9: 10, 10: 24})


def _judgement_rows(ws, start: int, rows: list[tuple[str, str, float]]) -> int:
    """Write judgement rows; return total row index."""
    for idx, (code, name, mx) in enumerate(rows, start + 1):
        r = idx
        ws.cell(r, 1, code)
        ws.cell(r, 2, name)
        ws.cell(r, 3, mx)
        ws.cell(r, 7, f'=IF(COUNT(D{r}:F{r})=0,"",AVERAGE(D{r}:F{r}))')
        ws.cell(r, 8, f'=IF(G{r}="",0,G{r}*C{r}/3)')
        ws.cell(r, 9, f'=IF(COUNT(D{r}:F{r})<2,"",MAX(D{r}:F{r})-MIN(D{r}:F{r}))')
        for c in range(4, 7):
            ws.cell(r, c).number_format = "0.0"
    return start + 1 + len(rows)


def sheet_modul_b(wb: Workbook) -> None:
    ws = wb.create_sheet("Modul B")
    ws["A1"] = "MODUL B — Jurnal / Laporan Teknis (10 poin: J 8 + M 2)"
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"], ws["G3"] = "Tim No:", "Nama Tim:", "Kontingen:", "Deadline:"
    ws["A4"], ws["C4"] = "Juri 1:", "Juri 2:"
    ws["E4"], ws["G4"] = "Juri 3:", "Tepat Waktu (1/0):"

    headers = [
        "Kode", "Aspek", "Tipe", "Max", "J1", "J2", "J3",
        "Rata-rata", "Skor Akhir", "Selisih", "Catatan",
    ]
    start = 6
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    style_header_row(ws, start, len(headers))

    j_rows = [
        ("B-J1", "Kualitas Fabrikasi Rangka & Stabilitas", 1.5),
        ("B-J2", "Manajemen Pengkabelan (Wiring)", 1.5),
        ("B-J3", "Integrasi Sistem Penggerak", 1.5),
        ("B-J4", "Kalibrasi Sensor & Aktuator", 1.5),
        ("B-J5", "Optimasi Software-Hardware", 1.5),
        ("B-J6", "Kelengkapan Dokumentasi", 0.5),
    ]
    for idx, (code, name, mx) in enumerate(j_rows, start + 1):
        r = idx
        ws.cell(r, 1, code)
        ws.cell(r, 2, name)
        ws.cell(r, 3, "J")
        ws.cell(r, 4, mx)
        ws.cell(r, 8, f'=IF(COUNT(E{r}:G{r})=0,"",AVERAGE(E{r}:G{r}))')
        ws.cell(r, 9, f'=IF(H{r}="",0,H{r}*D{r}/3)')
        ws.cell(r, 10, f'=IF(COUNT(E{r}:G{r})<2,"",MAX(E{r}:G{r})-MIN(E{r}:G{r}))')

    j_total = start + 1 + len(j_rows)
    ws.cell(j_total, 2, "SUBTOTAL JUDGEMENT")
    ws.cell(j_total, 4, 8.0)
    ws.cell(j_total, 9, f"=SUM(I{start+1}:I{j_total-1})")
    ws.cell(j_total, 2).font = Font(bold=True)

    m_row = j_total + 1
    ws.cell(m_row, 1, "B-M1")
    ws.cell(m_row, 2, "Ketepatan Mengumpulkan sesuai Waktu")
    ws.cell(m_row, 3, "M")
    ws.cell(m_row, 4, 2.0)
    ws.cell(m_row, 9, f"=IF(G4=\"\",0,G4*D{m_row})")

    total_row = m_row + 1
    ws.cell(total_row, 2, "TOTAL MODUL B")
    ws.cell(total_row, 4, 10.0)
    ws.cell(total_row, 9, f"=I{j_total}+I{m_row}")
    ws.cell(total_row, 2).font = Font(bold=True)

    chk_start = total_row + 2
    ws.cell(chk_start, 1, "CHECKLIST KELENGKAPAN (1=ada)")
    ws.cell(chk_start, 1).font = Font(bold=True)
    checklist = [
        "Analisis Tugas & Strategi", "Desain Mekanik", "Diagram Blok", "Bill of Materials",
        "Dokumentasi Fabrikasi", "Manajemen Kelistrikan", "Penempatan Sensor",
        "Arsitektur Software", "Algoritma Navigasi", "Algoritma Visi", "Flowchart",
        "Data Kalibrasi", "Tuning Kontroler", "Log Troubleshooting",
    ]
    for i, item in enumerate(checklist, chk_start + 1):
        ws.cell(i, 1, f"C{i - chk_start}")
        ws.cell(i, 2, item)
        ws.cell(i, 3, "")

    set_widths(ws, {1: 8, 2: 36, 3: 6, 4: 8, 5: 8, 6: 8, 7: 8, 8: 10, 9: 10, 10: 8, 11: 20})


def sheet_modul_c(wb: Workbook) -> None:
    ws = wb.create_sheet("Modul C")
    ws["A1"] = "MODUL C — Pembuatan & Perakitan Robot (10 poin, Judgement)"
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"] = "Tim No:", "Nama Tim:", "Waktu Mulai:"
    ws["G3"], ws["I3"] = "Waktu Selesai:", "Durasi (jam):"
    ws["A4"], ws["C4"], ws["E4"] = "Juri 1:", "Juri 2:", "Juri 3:"

    headers = [
        "Sub", "Kriteria", "Max", "J1", "J2", "J3",
        "Rata-rata", "Skor Akhir", "Selisih", "Catatan",
    ]
    start = 6
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    style_header_row(ws, start, len(headers))

    rows = [
        ("C1", "Perakitan L-Channel sebagai Penyangga Utama Lift", 2.5),
        ("C2", "Integrasi Sistem Angkat / Gripper / Lifter", 2.5),
        ("C3", "Kualitas Rewiring (rapi, aman)", 2.5),
        ("C4", "Kelayakan Operasional (STOP, struktur stabil)", 2.5),
    ]
    total_row = _judgement_rows(ws, start, rows)
    ws.cell(total_row, 2, "TOTAL MODUL C")
    ws.cell(total_row, 3, 10.0)
    ws.cell(total_row, 8, f"=SUM(H{start+1}:H{total_row-1})")
    ws.cell(total_row, 2).font = Font(bold=True)

    chk = total_row + 2
    ws.cell(chk, 1, "CHECKLIST INSPEKSI (1=Lulus 0=Tidak)")
    ws.cell(chk, 1).font = Font(bold=True)
    items = [
        "L-Channel sebagai penyangga utama angkat",
        "Sistem objek terpasang kuat di L-Channel",
        "Tidak ada kabel terbuka/terjepit",
        "Tombol STOP merah berfungsi",
        "Selesai dalam batas 2 jam",
        "Robot aman dioperasikan (K3)",
    ]
    for i, item in enumerate(items, chk + 1):
        ws.cell(i, 1, f"I{i-chk}")
        ws.cell(i, 2, item)

    set_widths(ws, {1: 6, 2: 42, 3: 8, 4: 8, 5: 8, 6: 8, 7: 10, 8: 10, 9: 8, 10: 22})


def sheet_modul_d(wb: Workbook) -> None:
    ws = wb.create_sheet("Modul D")
    ws["A1"] = "MODUL D — Gerakan Dasar & Manajemen Objek (12 poin, Measurement)"
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"] = "Tim No:", "Nama Tim:", "Tanggal:"
    ws["A4"] = "Juri Penilai:"

    headers = ["ID", "Evaluasi", "Indikator", "Bobot", "Hasil (1/0)", "Skor", "Catatan"]
    start = 6
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    style_header_row(ws, start, len(headers))

    items = [
        ("D1a", "Gerak maju 200-300 mm", "Observasi", 0.5),
        ("D1b", "Gerak mundur 200-300 mm", "Observasi", 0.5),
        ("D1c", "Gerak kanan 200-300 mm", "Observasi", 0.5),
        ("D1d", "Gerak kiri 200-300 mm", "Observasi", 0.5),
        ("D2a", "Maju stop 50-100 mm dari dinding", "Observasi", 0.3),
        ("D2b", "Kanan stop 50-100 mm dari dinding", "Observasi", 0.6),
        ("D3", "Ikuti garis U", "Line follower", 1.0),
        ("D4", "Deteksi merah → teks merah", "Monitor", 1.0),
        ("D5", "Deteksi hijau → teks hijau", "Monitor", 1.0),
        ("D6", "Deteksi biru → teks biru", "Monitor", 1.0),
        ("D7", "Deteksi O-merah", "Monitor", 1.0),
        ("D8", "Deteksi O-hijau", "Monitor", 1.0),
        ("D9", "Pick kubus dari rak otonom", "Observasi arm", 2.0),
        ("D10", "Place kubus ke standbox", "Observasi arm", 2.0),
    ]
    for i, (id_, ev, ind, bobot) in enumerate(items, start + 1):
        ws.cell(i, 1, id_)
        ws.cell(i, 2, ev)
        ws.cell(i, 3, ind)
        ws.cell(i, 4, bobot)
        ws.cell(i, 6, f"=IF(E{i}=\"\",0,E{i}*D{i})")

    tr = start + 1 + len(items)
    ws.cell(tr, 2, "TOTAL (skala 12)")
    ws.cell(tr, 4, 12.0)
    ws.cell(tr, 6, f"=ROUND(SUM(F{start+1}:F{tr-1})*12/{MODUL_D_BOBOT_SUM},2)")
    ws.cell(tr, 7, f"Bobot mentah item = {MODUL_D_BOBOT_SUM}")
    ws.cell(tr, 2).font = Font(bold=True)

    set_widths(ws, {1: 8, 2: 34, 3: 16, 4: 8, 5: 12, 6: 10, 7: 24})


def sheet_modul_e(wb: Workbook) -> None:
    ws = wb.create_sheet("Modul E")
    ws["A1"] = "MODUL E — Performansi Otonom (60 poin, Measurement)"
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"], ws["G3"] = "Tim No:", "Run ID:", "Hari:", "Trial/Final:"
    ws["A4"], ws["C4"] = "Arena No:", "Waktu START:"

    headers = [
        "No", "ID Kubus", "Warna", "Bentuk", "Posisi Awal", "Target Rak",
        "Target Marker", "Hasil (1/0)", "Bobot", "Skor", "Catatan",
    ]
    start = 6
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    style_header_row(ws, start, len(headers))

    for n in range(1, 10):
        r = start + n
        undian_row = UNDIAN_POS_FIRST + n - 1
        ws.cell(r, 1, n)
        ws.cell(r, 2, f"=IF({UNDIAN_SHEET}!A{undian_row}=\"\",\"\",{UNDIAN_SHEET}!A{undian_row})")
        ws.cell(r, 3, f'=IF(B{r}="","",VLOOKUP(B{r},{CAT_RANGE},2,FALSE))')
        ws.cell(r, 4, f'=IF(B{r}="","",VLOOKUP(B{r},{CAT_RANGE},3,FALSE))')
        ws.cell(r, 5, f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},{POS_RANGE},4,FALSE),""))')
        ws.cell(r, 9, "=60/9")
        ws.cell(r, 10, f"=IF(H{r}=\"\",0,H{r}*I{r})")

    tr = start + 10
    ws.cell(tr, 5, "TOTAL MODUL E")
    ws.cell(tr, 9, 60.0)
    ws.cell(tr, 10, f"=ROUND(SUM(J{start+1}:J{tr-1}),2)")
    ws.cell(tr, 5).font = Font(bold=True)

    ws.cell(tr + 2, 1, "Pelanggaran (1=ya):")
    pelanggaran = [
        "Sentuh laptop setelah SIAP",
        "Sentuh robot setelah START",
        "Tidak start via tombol robot",
        "Place manual oleh peserta",
    ]
    for i, p in enumerate(pelanggaran, tr + 3):
        ws.cell(i, 1, p)
        ws.cell(i, 2, 0)

    set_widths(ws, {1: 5, 2: 10, 3: 10, 4: 10, 5: 14, 6: 12, 7: 14, 8: 12, 9: 8, 10: 8, 11: 20})


def sheet_rekap(wb: Workbook) -> None:
    ws = wb.create_sheet("Rekapitulasi")
    ws["A1"] = "REKAPITULASI SKOR — Robot Bergerak Otonom LKS 2026"
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"], ws["C3"], ws["E3"] = "Tim No:", "Nama Tim:", "Kontingen:"

    headers = ["Modul", "Kriteria", "Bobot Max", "Skor", "Persen", "Metode"]
    start = 5
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    style_header_row(ws, start, len(headers))

    data = [
        ("A", "Organisasi & Manajemen Kerja", 8, "=\'Modul A\'!H11", "J"),
        ("B", "Jurnal Teknis", 10, "=\'Modul B\'!I15", "J+M"),
        ("C", "Perakitan Robot", 10, "=\'Modul C\'!H11", "J"),
        ("D", "Gerakan Dasar", 12, "=\'Modul D\'!F21", "M"),
        ("E", "Performa Otonom", 60, "=\'Modul E\'!J16", "M"),
    ]
    for i, (mod, crit, mx, formula, met) in enumerate(data, start + 1):
        ws.cell(i, 1, mod)
        ws.cell(i, 2, crit)
        ws.cell(i, 3, mx)
        if isinstance(formula, str) and formula.startswith("="):
            ws.cell(i, 4, formula)
        else:
            ws.cell(i, 4, formula)
        ws.cell(i, 5, f"=IF(C{i}=0,0,D{i}/C{i}*100)")
        ws.cell(i, 6, met)

    tr = start + 1 + len(data)
    ws.cell(tr, 2, "TOTAL")
    ws.cell(tr, 3, 100)
    ws.cell(tr, 4, f"=SUM(D{start+1}:D{tr-1})")
    ws.cell(tr, 5, f"=D{tr}")
    ws.cell(tr, 2).font = Font(bold=True)

    ws.cell(tr + 2, 1, "Skor CIS (0-100):")
    ws.cell(tr + 3, 1, "Skor Konversi (700):")
    ws.cell(tr + 4, 1, "Medallion (>700):")

    set_widths(ws, {1: 8, 2: 32, 3: 12, 4: 12, 5: 10, 6: 10})


def sheet_panduan(wb: Workbook) -> None:
    ws = wb.create_sheet("Panduan", 0)
    ws["A1"] = "PANDUAN PENGGUNAAN — LKS 2026 Penilaian Juri"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "SATU FILE INI untuk seluruh penilaian — jangan buka file CSV terpisah.",
        "",
        "1. Duplikat workbook ini per TIM (Save As).",
        "2. Modul A/B/C: isi skor Juri 1-3 (0-3). Skor otomatis.",
        "3. Modul B: isi Tepat Waktu (1/0) di G4.",
        "4. Undian Kubus: isi ID kubus setelah pengacakan — Warna/Bentuk otomatis.",
        "5. Log Run Otonom: checklist SIAP/START + kronologi; skor tarik dari Modul E.",
        "6. Modul D/E: isi Hasil 1 atau 0. Modul E: isi ID Kubus (kolom B) dari undian.",
        "7. Rekapitulasi: total otomatis dari Modul A-E.",
        "8. Input skor resmi ke CIS (Marking Scheme Chief Expert).",
        "",
        "Sheet: Panduan | Modul A-E | Undian Kubus | Log Run | Rekapitulasi",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 80


def main() -> None:
    wb = Workbook()
    sheet_panduan(wb)
    sheet_modul_a(wb)
    sheet_modul_b(wb)
    sheet_modul_c(wb)
    sheet_modul_d(wb)
    sheet_undian_kubus(wb)  # sebelum Modul E (referensi rumus)
    sheet_modul_e(wb)
    sheet_log_run(wb)
    sheet_rekap(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
