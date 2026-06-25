"""Generate SMK Muhammadiyah variant — Job Arena tetap + Modul E manual (tanpa Undian otomatis)."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import generate_excel_templates as gen

ROOT = Path(__file__).resolve().parent.parent
OUT_XLSX = ROOT / "templates" / "LKS2026-Penilaian-Juri TIM SMK Muhammadiyah.xlsx"
OUT_XLSM = OUT_XLSX.with_suffix(".xlsm")
JOB_IMAGE_SRC = Path(r"C:\Users\anton\Downloads\Job-SMK Muhammadyah_page_1.png")
JOB_IMAGE_REPO = ROOT / "assets" / "job-muhammadiyah-arena.png"
GITHUB_JOB = f"{gen.GITHUB_REPO}/blob/main/docs/PANDUAN-JOB-MUHAMMADIYAH.md"

# Posisi tetap sesuai denah Job SMK Muhammadiyah (KS1–KS9)
JOB_KUBUS = [
  # id, warna, bentuk, stand, target_rak, slot_warna
    ("K-M1", "Merah", "Lingkaran", "KS7", "Rak 1 (Lingkaran)", "Merah"),
    ("K-M2", "Merah", "Persegi", "KS4", "Rak 2 (Persegi)", "Merah"),
    ("K-M3", "Merah", "Segitiga", "KS1", "Rak 3 (Segitiga)", "Merah"),
    ("K-H1", "Hijau", "Lingkaran", "KS8", "Rak 1 (Lingkaran)", "Hijau"),
    ("K-H2", "Hijau", "Persegi", "KS5", "Rak 2 (Persegi)", "Hijau"),
    ("K-H3", "Hijau", "Segitiga", "KS2", "Rak 3 (Segitiga)", "Hijau"),
    ("K-B1", "Biru", "Lingkaran", "KS9", "Rak 1 (Lingkaran)", "Biru"),
    ("K-B2", "Biru", "Persegi", "KS6", "Rak 2 (Persegi)", "Biru"),
    ("K-B3", "Biru", "Segitiga", "KS3", "Rak 3 (Segitiga)", "Biru"),
]

JOB_RULES = [
    "Memindahkan semua kubus ke rak yang sesuai bentuk & warna (Lingkaran→Rak1, Persegi→Rak2, Segitiga→Rak3).",
    "Saat membawa kubus: robot BOLEH melewati Gate.",
    "Saat TIDAK membawa kubus: robot WAJIB melewati Obstacle.",
]

MODUL_E_START = 7
MODUL_E_TOTAL_ROW = 16


def _ensure_job_image() -> Path | None:
    if JOB_IMAGE_REPO.exists():
        return JOB_IMAGE_REPO
    if JOB_IMAGE_SRC.exists():
        JOB_IMAGE_REPO.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(JOB_IMAGE_SRC, JOB_IMAGE_REPO)
        return JOB_IMAGE_REPO
    return None


def sheet_panduan_muhammadiyah(wb: Workbook) -> None:
    ws = wb.create_sheet("Panduan", 0)
    ws["A1"] = "PANDUAN — TIM SMK MUHAMMADIYAH (Job Tetap, Modul E Manual)"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "VERSI INI untuk latihan / job internal Muhammadiyah — BUKAN template undian resmi LKS nasional.",
        "",
        "1. Baca sheet JOB ARENA — denah KS1–KS9, rak, Gate, Obstacle.",
        "2. Modul E: kolom B–G sudah terisi (posisi & target). Juri isi H (Hasil Run 1/0) saja.",
        "3. Log Run Otonom: stopwatch SIAP/START/SELESAI (.xlsm) atau Google Sheets Apps Script.",
        "4. Modul A–D & Rekapitulasi sama seperti template utama.",
        "",
        "Sheet: Panduan | Job Arena | Modul A–E | Log Run | Rekapitulasi",
        "",
        f"Panduan lengkap: {GITHUB_JOB}",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(i, 1, line)
        if line.startswith("http") or "github" in line:
            ws.cell(i, 1).hyperlink = GITHUB_JOB
            ws.cell(i, 1).font = Font(color="0563C1", underline="single")
    ws.column_dimensions["A"].width = 90


def sheet_job_arena(wb: Workbook) -> None:
    ws = wb.create_sheet("Job Arena")
    gen.sheet_title(ws, "JOB ARENA — SMK MUHAMMADIYAH (layout tetap)", merge_end="L1")
    gen.add_github_help(
        ws,
        GITHUB_JOB,
        "Panduan job Muhammadiyah di GitHub (klik di sini)",
        merge_end="L",
    )

    img_path = _ensure_job_image()
    row_after_img = 3
    if img_path:
        try:
            img = XLImage(str(img_path))
            img.width = min(img.width, 720)
            img.height = min(img.height, 400)
            ws.add_image(img, "A3")
            row_after_img = 28
        except Exception:
            row_after_img = 3

    r = row_after_img
    ws.cell(r, 1, "KATALOG KUBUS").font = Font(bold=True, size=12)
    r += 1
    for c, h in enumerate(["ID Kubus", "Warna", "Bentuk", "Stand (KS)", "Target Rak", "Slot warna"], 1):
        ws.cell(r, c, h)
    gen.style_header_row(ws, r, 6)
    for row in JOB_KUBUS:
        r += 1
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    r += 2
    ws.cell(r, 1, "POSISI STAND (KS) — denah").font = Font(bold=True, size=12)
    r += 1
    stands = [
        ("KS1", "Kiri arena", "K-M3"),
        ("KS2", "Kiri arena", "K-H3"),
        ("KS3", "Kiri arena", "K-B3"),
        ("KS4", "Atas arena", "K-M2"),
        ("KS5", "Atas arena", "K-H2"),
        ("KS6", "Atas arena", "K-B2"),
        ("KS7", "Kanan arena", "K-M1"),
        ("KS8", "Kanan arena", "K-H1"),
        ("KS9", "Kanan arena", "K-B1"),
    ]
    for c, h in enumerate(["Stand", "Lokasi", "Kubus awal"], 1):
        ws.cell(r, c, h)
    gen.style_header_row(ws, r, 3)
    for stand, loc, kid in stands:
        r += 1
        ws.cell(r, 1, stand)
        ws.cell(r, 2, loc)
        ws.cell(r, 3, kid)

    r += 2
    ws.cell(r, 1, "RAK TARGET").font = Font(bold=True, size=12)
    r += 1
    for c, h in enumerate(["Rak", "Bentuk", "Marker slot (M-H-B)"], 1):
        ws.cell(r, c, h)
    gen.style_header_row(ws, r, 3)
    for rak, bentuk in [("Rak 1", "Lingkaran"), ("Rak 2", "Persegi"), ("Rak 3", "Segitiga")]:
        r += 1
        ws.cell(r, 1, rak)
        ws.cell(r, 2, bentuk)
        ws.cell(r, 3, "Merah | Hijau | Biru (kiri–tengah–kanan)")

    r += 2
    ws.cell(r, 1, "JOB DESKRIPSI").font = Font(bold=True, size=12)
    for i, rule in enumerate(JOB_RULES, 1):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1, f"{i}. {rule}")
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)

    r += 2
    ws.cell(r, 1, "OBSTACLE & GATE (referensi denah)").font = Font(bold=True)
    r += 1
    ws.cell(r, 1, "Gate: kiri bawah — boleh dilalui saat bawa kubus.")
    r += 1
    ws.cell(r, 1, "Obstacle: 2 buah di arena — wajib dilalui saat tanpa kubus.")

    gen.set_widths(ws, {1: 14, 2: 12, 3: 12, 4: 10, 5: 18, 6: 14, 7: 12, 8: 12})


def sheet_modul_e_manual(wb: Workbook) -> None:
    ws = wb.create_sheet("Modul E")
    ws["A1"] = "MODUL E — Job Tetap SMK Muhammadiyah (60 poin, manual)"
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(bold=True, size=14)
    gen.add_github_help(ws, GITHUB_JOB, "Panduan job Muhammadiyah (GitHub)", merge_end="K")

    ws["A3"], ws["C3"], ws["E3"], ws["G3"] = "Tim No:", "Run ID:", "Hari:", "Trial/Final:"
    ws["A4"], ws["C4"] = "Arena No:", "Waktu START:"
    ws["D4"] = "='Log Run Otonom'!E8"
    ws["D4"].number_format = "hh:mm:ss"

    ws.merge_cells("A5:K5")
    hint = ws["A5"]
    hint.value = (
        "JOB TETAP — tidak perlu Undian Kubus. Kolom B–G sudah terisi dari sheet Job Arena. "
        "Juri isi kolom H (Hasil Run 1/0) setelah run. Sukses = kubus di rak bentuk benar + slot warna cocok + stabil."
    )
    hint.font = Font(italic=True, color="1F4E79")
    hint.fill = PatternFill("solid", fgColor="E8F4FC")
    hint.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[5].height = 36

    headers = [
        "No", "ID Kubus", "Warna", "Bentuk", "Posisi Awal (KS)",
        "Target Rak", "Slot Target", "Hasil Run (1/0)", "Bobot", "Skor", "Catatan",
    ]
    hdr = MODUL_E_START - 1
    for i, h in enumerate(headers, 1):
        ws.cell(hdr, i, h)
    gen.style_header_row(ws, hdr, len(headers))

    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv)

    for n, row in enumerate(JOB_KUBUS, 1):
        r = MODUL_E_START + n - 1
        kid, warna, bentuk, ks, rak, slot = row
        ws.cell(r, 1, n)
        ws.cell(r, 2, kid)
        ws.cell(r, 3, warna)
        ws.cell(r, 4, bentuk)
        ws.cell(r, 5, ks)
        ws.cell(r, 6, rak)
        ws.cell(r, 7, slot)
        ws.cell(r, 9, "=60/9")
        ws.cell(r, 10, f'=IF(H{r}="",0,H{r}*I{r})')
        dv.add(ws.cell(r, 8))

    tr = MODUL_E_TOTAL_ROW
    ws.cell(tr, 6, "TOTAL MODUL E")
    ws.cell(tr, 9, 60.0)
    ws.cell(tr, 10, f"=ROUND(SUM(J{MODUL_E_START}:J{tr-1}),2)")
    ws.cell(tr, 6).font = Font(bold=True)

    ws.cell(tr + 2, 1, "Pelanggaran (1=ya):")
    for i, p in enumerate(
        [
            "Sentuh laptop setelah SIAP",
            "Sentuh robot setelah START",
            "Tidak start via tombol robot",
            "Place manual oleh peserta",
        ],
        tr + 3,
    ):
        ws.cell(i, 1, p)
        ws.cell(i, 2, 0)

    gen.set_widths(ws, {1: 5, 2: 10, 3: 10, 4: 10, 5: 12, 6: 16, 7: 12, 8: 14, 9: 8, 10: 8, 11: 18})


def build_workbook() -> Workbook:
    wb = Workbook()
    sheet_panduan_muhammadiyah(wb)
    gen.sheet_modul_a(wb)
    gen.sheet_modul_b(wb)
    gen.sheet_modul_c(wb)
    gen.sheet_modul_d(wb)
    sheet_job_arena(wb)
    sheet_modul_e_manual(wb)
    gen.sheet_log_run(wb)
    gen.sheet_rekap(wb)
    return wb


def main() -> None:
    wb = build_workbook()
    out_xlsx = OUT_XLSX
    out_xlsm = OUT_XLSM
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_xlsx.with_name("_tmp_muhammadiyah.xlsx")
    wb.save(tmp)
    try:
        tmp.replace(out_xlsx)
        print(f"Created: {out_xlsx}")
    except PermissionError:
        print(f"WARN: file terkunci — simpan ke {tmp}")
        out_xlsx = tmp

    try:
        from add_log_run_stopwatch import apply_stopwatch

        apply_stopwatch(out_xlsx, out_xlsm)
        print(f"Created: {out_xlsm} (stopwatch)")
    except Exception as exc:
        print(f"WARN: .xlsm tidak dibuat — {exc}")


if __name__ == "__main__":
    main()
